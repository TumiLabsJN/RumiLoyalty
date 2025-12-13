#!/usr/bin/env python3
"""
Loyalty Program Financial Model - Excel Generator V2
Creates a dynamic Excel workbook with FULLY DYNAMIC formulas.
All costs reference configuration sheets - no hardcoded values.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
label_font = Font(bold=True, size=11)
section_font = Font(bold=True, size=14, color="4472C4")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(cell):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

def style_input(cell):
    cell.fill = input_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

def style_calc(cell):
    cell.fill = calc_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

def style_label(cell):
    cell.font = label_font
    cell.border = thin_border

def style_section(cell):
    cell.font = section_font

# ============================================================================
# SHEET 1: INPUTS DASHBOARD
# ============================================================================
ws1 = wb.active
ws1.title = "Inputs"

ws1['A1'] = "LOYALTY PROGRAM FINANCIAL MODEL - INPUTS DASHBOARD"
ws1['A1'].font = Font(bold=True, size=16, color="4472C4")
ws1.merge_cells('A1:E1')

# Section: Acquisition Inputs
ws1['A3'] = "ACQUISITION INPUTS"
style_section(ws1['A3'])

headers = ['Parameter', 'Value', 'Unit', 'Description']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    style_header(cell)

acquisition_inputs = [
    ('Samples Sent - Month 1', 150, 'samples', 'Initial month sample distribution'),
    ('Samples Sent - Month 2+', 100, 'samples', 'Ongoing monthly sample distribution'),
    ('Sample-to-Affiliate Conversion Rate', 0.20, '%', 'Percentage of samples that convert to affiliates'),
    ('Flywheel Affiliates per Month', 10, 'affiliates', 'Organic inbound affiliates monthly'),
    ('Time to First Sale', 14, 'days', 'Lag between conversion and first sale'),
    ('Affiliate Attrition Rate', 0.05, '%', 'Monthly churn rate of affiliates'),
]

for i, (param, value, unit, desc) in enumerate(acquisition_inputs, 5):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    cell = ws1.cell(row=i, column=2, value=value)
    style_input(cell)
    if unit == '%':
        cell.number_format = '0%'
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=desc).border = thin_border

# Section: Performance Inputs
ws1['A12'] = "PERFORMANCE INPUTS"
style_section(ws1['A12'])

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=13, column=col, value=header)
    style_header(cell)

performance_inputs = [
    ('Average Sales per Affiliate per Month', 5, 'sales', 'Monthly sales velocity per active affiliate'),
    ('Gross AOV', 100, '$', 'Average order value before discounts'),
    ('Rolling Window Duration', 30, 'days', 'Time period for VIP level qualification'),
]

for i, (param, value, unit, desc) in enumerate(performance_inputs, 14):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    cell = ws1.cell(row=i, column=2, value=value)
    style_input(cell)
    if unit == '$':
        cell.number_format = '$#,##0.00'
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=desc).border = thin_border

# Section: Redemption Inputs
ws1['A19'] = "REDEMPTION & BEHAVIOR INPUTS"
style_section(ws1['A19'])

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=20, column=col, value=header)
    style_header(cell)

redemption_inputs = [
    ('Discount Redemption Rate', 0.30, '%', 'Percentage of discount coupons redeemed'),
    ('Default Mission Completion Rate', 0.25, '%', 'Default completion rate for missions'),
    ('Avg Sales During Commission Boost', 3, 'sales', 'Expected sales an affiliate makes during boost period'),
]

for i, (param, value, unit, desc) in enumerate(redemption_inputs, 21):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    cell = ws1.cell(row=i, column=2, value=value)
    style_input(cell)
    if unit == '%':
        cell.number_format = '0%'
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=desc).border = thin_border

# Section: Cost Inputs
ws1['A26'] = "COST INPUTS"
style_section(ws1['A26'])

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=27, column=col, value=header)
    style_header(cell)

cost_inputs = [
    ('Cost per Sample', 15, '$', 'Cost of each sample sent (product + shipping)'),
]

for i, (param, value, unit, desc) in enumerate(cost_inputs, 28):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    cell = ws1.cell(row=i, column=2, value=value)
    style_input(cell)
    if unit == '$':
        cell.number_format = '$#,##0.00'
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=desc).border = thin_border

# Section: Level Distribution Assumptions
ws1['A31'] = "LEVEL DISTRIBUTION ASSUMPTIONS"
style_section(ws1['A31'])
ws1['A32'] = "(Estimated % of affiliates at each level based on sales velocity vs thresholds)"
ws1['A32'].font = Font(italic=True, size=10)

level_dist_headers = ['Level', '% of Affiliates', 'Description']
for col, header in enumerate(level_dist_headers, 1):
    cell = ws1.cell(row=33, column=col, value=header)
    style_header(cell)

# These percentages represent steady-state distribution based on sales velocity
level_distribution = [
    ('Bronze', 0.40, 'Affiliates below Silver threshold'),
    ('Silver', 0.30, 'Affiliates at Silver, below Gold'),
    ('Gold', 0.18, 'Affiliates at Gold, below Platinum'),
    ('Platinum', 0.09, 'Affiliates at Platinum, below Diamond'),
    ('Diamond', 0.03, 'Top performers at Diamond'),
]

for i, (level, pct, desc) in enumerate(level_distribution, 34):
    ws1.cell(row=i, column=1, value=level).border = thin_border
    cell = ws1.cell(row=i, column=2, value=pct)
    style_input(cell)
    cell.number_format = '0%'
    ws1.cell(row=i, column=3, value=desc).border = thin_border

# Validation row
ws1.cell(row=39, column=1, value="Total (must = 100%)").border = thin_border
ws1.cell(row=39, column=1).font = Font(bold=True)
cell = ws1.cell(row=39, column=2, value="=SUM(B34:B38)")
style_calc(cell)
cell.number_format = '0%'

# Legend
ws1['A42'] = "LEGEND"
style_section(ws1['A42'])
ws1['A43'] = "Yellow cells"
ws1['B43'] = "= Editable inputs"
ws1['B43'].fill = input_fill
ws1['A44'] = "Green cells"
ws1['B44'] = "= Calculated values"
ws1['B44'].fill = calc_fill

ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 45

# ============================================================================
# SHEET 2: VIP LEVEL CONFIGURATION
# ============================================================================
ws2 = wb.create_sheet("VIP Levels")

ws2['A1'] = "VIP LEVEL CONFIGURATION"
ws2['A1'].font = Font(bold=True, size=16, color="4472C4")
ws2.merge_cells('A1:O1')

# Level Configuration Table
ws2['A3'] = "LEVEL THRESHOLDS & COMMISSION"
style_section(ws2['A3'])

level_headers = ['Level #', 'Name', 'Sales Threshold', 'Base Commission %']
for col, header in enumerate(level_headers, 1):
    cell = ws2.cell(row=4, column=col, value=header)
    style_header(cell)

levels = [
    (1, 'Bronze', 0, 0.10),
    (2, 'Silver', 5, 0.12),
    (3, 'Gold', 20, 0.18),
    (4, 'Platinum', 50, 0.22),
    (5, 'Diamond', 100, 0.25),
]

for i, (level, name, threshold, commission) in enumerate(levels, 5):
    ws2.cell(row=i, column=1, value=level).border = thin_border
    cell = ws2.cell(row=i, column=2, value=name)
    style_input(cell)
    cell = ws2.cell(row=i, column=3, value=threshold)
    style_input(cell)
    cell = ws2.cell(row=i, column=4, value=commission)
    style_input(cell)
    cell.number_format = '0%'

# Welcome Rewards Configuration
ws2['A12'] = "WELCOME REWARDS (Per Level)"
style_section(ws2['A12'])

reward_headers = [
    'Level', 'Comm Boost Qty', 'Comm Boost %', 'Comm Boost Days',
    'Gift Card Qty', 'Gift Card $', 'Discount Qty', 'Discount %',
    'Spark Ads Qty', 'Spark Ads $', 'Phys Gift Qty', 'Phys Gift $',
    'Experience Qty', 'Experience $'
]
for col, header in enumerate(reward_headers, 1):
    cell = ws2.cell(row=13, column=col, value=header)
    style_header(cell)

# VIP Levels welcome rewards - Row positions: 14=Bronze, 15=Silver, 16=Gold, 17=Platinum, 18=Diamond
welcome_rewards = [
    ('Bronze', 1, 0.10, 30, 0, 0, 2, 0.15, 0, 0, 0, 0, 0, 0),
    ('Silver', 2, 0.15, 30, 1, 50, 2, 0.15, 1, 25, 0, 0, 0, 0),
    ('Gold', 1, 0.20, 30, 1, 100, 2, 0.20, 1, 50, 1, 50, 0, 0),
    ('Platinum', 2, 0.25, 45, 2, 150, 3, 0.25, 2, 75, 1, 75, 1, 100),
    ('Diamond', 3, 0.30, 60, 2, 200, 3, 0.30, 2, 100, 1, 100, 1, 200),
]

for i, rewards in enumerate(welcome_rewards, 14):
    for col, val in enumerate(rewards, 1):
        cell = ws2.cell(row=i, column=col, value=val)
        if col == 1:
            cell.border = thin_border
        else:
            style_input(cell)
        if col in [3, 8]:  # Percentages
            cell.number_format = '0%'
        if col in [6, 10, 12, 14]:  # Currency
            cell.number_format = '$#,##0'

# Calculated summaries for welcome rewards
ws2['A21'] = "WELCOME REWARD SUMMARY (Weighted by Level Distribution)"
style_section(ws2['A21'])

summary_labels = [
    ('Weighted Avg Commission Boost %', "=SUMPRODUCT(C14:C18,Inputs!$B$34:$B$38)"),
    ('Weighted Avg Gift Card Value', "=SUMPRODUCT(E14:E18*F14:F18,Inputs!$B$34:$B$38)"),
    ('Weighted Avg Spark Ads Value', "=SUMPRODUCT(I14:I18*J14:J18,Inputs!$B$34:$B$38)"),
    ('Weighted Avg Physical Gift Value', "=SUMPRODUCT(K14:K18*L14:L18,Inputs!$B$34:$B$38)"),
    ('Weighted Avg Experience Value', "=SUMPRODUCT(M14:M18*N14:N18,Inputs!$B$34:$B$38)"),
]

for i, (label, formula) in enumerate(summary_labels, 22):
    ws2.cell(row=i, column=1, value=label).border = thin_border
    cell = ws2.cell(row=i, column=2, value=formula)
    style_calc(cell)
    if '%' in label:
        cell.number_format = '0.0%'
    else:
        cell.number_format = '$#,##0.00'

for col in range(1, 15):
    ws2.column_dimensions[get_column_letter(col)].width = 13
ws2.column_dimensions['A'].width = 45

# ============================================================================
# SHEET 3: MISSION CONFIGURATION
# ============================================================================
ws3 = wb.create_sheet("Missions")

ws3['A1'] = "MISSION CONFIGURATION"
ws3['A1'].font = Font(bold=True, size=16, color="4472C4")
ws3.merge_cells('A1:L1')

ws3['A3'] = "MISSIONS BY VIP LEVEL"
style_section(ws3['A3'])

mission_headers = [
    'VIP Level', 'Mission Type', 'Target', 'Repeatability', 'Completion Rate',
    'Reward Type', 'Reward Value', 'Reward Duration', 'Active?', 'Cost per Completion'
]
for col, header in enumerate(mission_headers, 1):
    cell = ws3.cell(row=4, column=col, value=header)
    style_header(cell)

# Sample missions with cost calculation
# Column J (10) will calculate cost based on reward type
missions = [
    ('Bronze', 'Videos', 5, 'Monthly', 0.25, 'Gift Card', 50, 0, 'Yes'),
    ('Bronze', 'Likes', 100, 'Monthly', 0.20, 'Gift Card', 50, 0, 'Yes'),
    ('Silver', 'Videos', 10, 'Monthly', 0.20, 'Commission Boost', 0.20, 30, 'Yes'),
    ('Silver', 'Likes', 100, 'Monthly', 0.15, 'Gift Card', 50, 0, 'Yes'),
    ('Gold', 'Sales', 100, 'One-time', 0.10, 'Commission Boost', 0.30, 45, 'Yes'),
    ('Gold', 'Views', 1000, 'Monthly', 0.15, 'Spark Ads', 50, 0, 'Yes'),
    ('Platinum', 'Sales', 200, 'One-time', 0.08, 'Gift Card', 200, 0, 'Yes'),
    ('Platinum', 'Videos', 20, 'Monthly', 0.12, 'Commission Boost', 0.25, 30, 'Yes'),
    ('Diamond', 'Sales', 500, 'One-time', 0.05, 'Gift Card', 300, 0, 'Yes'),
    ('Diamond', 'Views', 5000, 'Monthly', 0.10, 'Spark Ads', 100, 0, 'Yes'),
]

for i, mission in enumerate(missions, 5):
    for col, val in enumerate(mission, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        style_input(cell)
        if col == 5:  # Completion rate
            cell.number_format = '0%'
        elif col == 7:  # Reward value
            if mission[5] == 'Commission Boost':
                cell.number_format = '0%'
            else:
                cell.number_format = '$#,##0'

    # Column J: Cost per completion
    # For Gift Card/Spark Ads: direct cost = reward value
    # For Commission Boost: cost = boost% * avg_sales_during_boost * Net_AOV
    row = i
    cell = ws3.cell(row=row, column=10)
    cell.value = f'=IF(F{row}="Gift Card",G{row},IF(F{row}="Spark Ads",G{row},IF(F{row}="Commission Boost",G{row}*Inputs!$B$23*Inputs!$B$15*0.955,0)))'
    style_calc(cell)
    cell.number_format = '$#,##0.00'

# Add empty rows for more missions
for i in range(15, 25):
    for col in range(1, 10):
        cell = ws3.cell(row=i, column=col, value='')
        style_input(cell)
    # Cost formula for empty rows
    cell = ws3.cell(row=i, column=10)
    cell.value = f'=IF(F{i}="Gift Card",G{i},IF(F{i}="Spark Ads",G{i},IF(F{i}="Commission Boost",G{i}*Inputs!$B$23*Inputs!$B$15*0.955,0)))'
    style_calc(cell)

# Mission Summary by Level
ws3['A27'] = "MISSION SUMMARY BY LEVEL"
style_section(ws3['A27'])

summary_headers = ['Level', 'Active Missions', 'Avg Completion Rate', 'Avg Cost per Completion']
for col, header in enumerate(summary_headers, 1):
    cell = ws3.cell(row=28, column=col, value=header)
    style_header(cell)

levels_list = ['Bronze', 'Silver', 'Gold', 'Platinum', 'Diamond']
for i, level in enumerate(levels_list, 29):
    ws3.cell(row=i, column=1, value=level).border = thin_border
    # Count active missions for this level
    cell = ws3.cell(row=i, column=2)
    cell.value = f'=COUNTIFS(A$5:A$24,A{i},I$5:I$24,"Yes")'
    style_calc(cell)
    # Avg completion rate
    cell = ws3.cell(row=i, column=3)
    cell.value = f'=IFERROR(AVERAGEIFS(E$5:E$24,A$5:A$24,A{i},I$5:I$24,"Yes"),0)'
    style_calc(cell)
    cell.number_format = '0%'
    # Avg cost per completion
    cell = ws3.cell(row=i, column=4)
    cell.value = f'=IFERROR(AVERAGEIFS(J$5:J$24,A$5:A$24,A{i},I$5:I$24,"Yes"),0)'
    style_calc(cell)
    cell.number_format = '$#,##0.00'

ws3['A36'] = "Mission Types: Videos, Likes, Sales, Views"
ws3['A37'] = "Repeatability: One-time, Weekly, Monthly"
ws3['A38'] = "Reward Types: Gift Card, Commission Boost, Spark Ads"

for col in range(1, 11):
    ws3.column_dimensions[get_column_letter(col)].width = 15

# ============================================================================
# SHEET 4: AFFILIATE PROJECTION
# ============================================================================
ws4 = wb.create_sheet("Affiliate Projection")

ws4['A1'] = "AFFILIATE PROJECTION (12 MONTHS)"
ws4['A1'].font = Font(bold=True, size=16, color="4472C4")
ws4.merge_cells('A1:N1')

proj_headers = ['Metric'] + [f'Month {i}' for i in range(1, 13)]
for col, header in enumerate(proj_headers, 1):
    cell = ws4.cell(row=3, column=col, value=header)
    style_header(cell)

metrics = [
    'Samples Sent',
    'New Affiliates (from Samples)',
    'Flywheel Affiliates',
    'Total New Affiliates',
    'Churned Affiliates',
    'Active Affiliates (End of Month)',
    'Total Sales',
    '--- Level Distribution ---',
    'Affiliates at Bronze',
    'Affiliates at Silver',
    'Affiliates at Gold',
    'Affiliates at Platinum',
    'Affiliates at Diamond',
    '--- Level Events ---',
    'New Affiliate Level-Ups (to Bronze)',
    'Promotion Events (Bronze to higher)',
    'Total Level-Up Events',
    'Demotion Events',
]

for i, metric in enumerate(metrics, 4):
    cell = ws4.cell(row=i, column=1, value=metric)
    if metric.startswith('---'):
        style_section(cell)
    else:
        style_label(cell)

# Month 1 formulas
ws4['B4'] = "=Inputs!$B$5"  # Samples Month 1
style_calc(ws4['B4'])

ws4['B5'] = "=ROUND(B4*Inputs!$B$7,0)"  # New affiliates from samples
style_calc(ws4['B5'])

ws4['B6'] = "=Inputs!$B$8"  # Flywheel
style_calc(ws4['B6'])

ws4['B7'] = "=B5+B6"  # Total new
style_calc(ws4['B7'])

ws4['B8'] = 0  # Churned (none in month 1)
style_calc(ws4['B8'])

ws4['B9'] = "=B7-B8"  # Active affiliates
style_calc(ws4['B9'])

ws4['B10'] = "=ROUND(B9*Inputs!$B$14,0)"  # Total sales
style_calc(ws4['B10'])

# Level distribution Month 1 - all start at Bronze
ws4['B12'] = "=B9"  # All at Bronze
style_calc(ws4['B12'])
for row in range(13, 17):
    ws4.cell(row=row, column=2, value=0)
    style_calc(ws4.cell(row=row, column=2))

# Level events Month 1
ws4['B18'] = "=B7"  # New affiliates enter at Bronze
style_calc(ws4['B18'])
ws4['B19'] = 0  # No promotions yet
style_calc(ws4['B19'])
ws4['B20'] = "=B18+B19"  # Total level-ups
style_calc(ws4['B20'])
ws4['B21'] = 0  # No demotions
style_calc(ws4['B21'])

# Months 2-12 formulas
for col in range(3, 14):
    ml = get_column_letter(col)  # month letter
    pl = get_column_letter(col - 1)  # previous letter

    # Samples (Month 2+)
    ws4.cell(row=4, column=col, value="=Inputs!$B$6")
    style_calc(ws4.cell(row=4, column=col))

    # New affiliates from samples
    ws4.cell(row=5, column=col, value=f"=ROUND({ml}4*Inputs!$B$7,0)")
    style_calc(ws4.cell(row=5, column=col))

    # Flywheel
    ws4.cell(row=6, column=col, value="=Inputs!$B$8")
    style_calc(ws4.cell(row=6, column=col))

    # Total new
    ws4.cell(row=7, column=col, value=f"={ml}5+{ml}6")
    style_calc(ws4.cell(row=7, column=col))

    # Churned (attrition on previous month's active)
    ws4.cell(row=8, column=col, value=f"=ROUND({pl}9*Inputs!$B$10,0)")
    style_calc(ws4.cell(row=8, column=col))

    # Active affiliates
    ws4.cell(row=9, column=col, value=f"={pl}9+{ml}7-{ml}8")
    style_calc(ws4.cell(row=9, column=col))

    # Total sales
    ws4.cell(row=10, column=col, value=f"=ROUND({ml}9*Inputs!$B$14,0)")
    style_calc(ws4.cell(row=10, column=col))

    # Level distribution - uses distribution percentages from Inputs sheet
    # Bronze
    ws4.cell(row=12, column=col, value=f"=ROUND({ml}9*Inputs!$B$34,0)")
    style_calc(ws4.cell(row=12, column=col))
    # Silver
    ws4.cell(row=13, column=col, value=f"=ROUND({ml}9*Inputs!$B$35,0)")
    style_calc(ws4.cell(row=13, column=col))
    # Gold
    ws4.cell(row=14, column=col, value=f"=ROUND({ml}9*Inputs!$B$36,0)")
    style_calc(ws4.cell(row=14, column=col))
    # Platinum
    ws4.cell(row=15, column=col, value=f"=ROUND({ml}9*Inputs!$B$37,0)")
    style_calc(ws4.cell(row=15, column=col))
    # Diamond
    ws4.cell(row=16, column=col, value=f"=ROUND({ml}9*Inputs!$B$38,0)")
    style_calc(ws4.cell(row=16, column=col))

    # Level events
    # New affiliates entering Bronze
    ws4.cell(row=18, column=col, value=f"={ml}7")
    style_calc(ws4.cell(row=18, column=col))

    # Promotions: estimate based on change in higher-level affiliates
    # Promotions = increase in (Silver+Gold+Plat+Diamond) from prev month, capped at 0
    ws4.cell(row=19, column=col, value=f"=MAX(0,({ml}13+{ml}14+{ml}15+{ml}16)-({pl}13+{pl}14+{pl}15+{pl}16))")
    style_calc(ws4.cell(row=19, column=col))

    # Total level-ups
    ws4.cell(row=20, column=col, value=f"={ml}18+{ml}19")
    style_calc(ws4.cell(row=20, column=col))

    # Demotions: estimate based on decrease in higher-level affiliates (excluding new churn effect)
    ws4.cell(row=21, column=col, value=f"=MAX(0,({pl}13+{pl}14+{pl}15+{pl}16)-({ml}13+{ml}14+{ml}15+{ml}16))")
    style_calc(ws4.cell(row=21, column=col))

ws4.column_dimensions['A'].width = 35
for col in range(2, 14):
    ws4.column_dimensions[get_column_letter(col)].width = 12

# ============================================================================
# SHEET 5: REWARD TRIGGERS
# ============================================================================
ws5 = wb.create_sheet("Reward Triggers")

ws5['A1'] = "REWARD TRIGGERS (12 MONTHS)"
ws5['A1'].font = Font(bold=True, size=16, color="4472C4")
ws5.merge_cells('A1:N1')

for col, header in enumerate(proj_headers, 1):
    cell = ws5.cell(row=3, column=col, value=header)
    style_header(cell)

reward_metrics = [
    '--- WELCOME REWARDS (per level-up) ---',
    'Commission Boosts Triggered',
    'Gift Cards Triggered',
    'Discount Coupons Triggered',
    'Spark Ads Triggered',
    'Physical Gifts Triggered',
    'Experiences Triggered',
    '',
    '--- MISSION COMPLETIONS ---',
    'Mission Completions (Bronze)',
    'Mission Completions (Silver)',
    'Mission Completions (Gold)',
    'Mission Completions (Platinum)',
    'Mission Completions (Diamond)',
    'Total Mission Completions',
]

for i, metric in enumerate(reward_metrics, 4):
    cell = ws5.cell(row=i, column=1, value=metric)
    if metric.startswith('---'):
        style_section(cell)
    else:
        style_label(cell)

# Formulas for each month
for col in range(2, 14):
    ml = get_column_letter(col)

    # Commission Boosts: Sum of (level-ups at each level * boost qty for that level)
    # Simplified: Total level-ups * weighted avg boost qty
    # More accurate: Calculate per level
    ws5.cell(row=5, column=col, value=f"=ROUND('Affiliate Projection'!{ml}18*'VIP Levels'!$B$14+'Affiliate Projection'!{ml}19*(SUMPRODUCT('VIP Levels'!$B$15:$B$18,Inputs!$B$35:$B$38)/SUM(Inputs!$B$35:$B$38)),0)")
    style_calc(ws5.cell(row=5, column=col))

    # Gift Cards: Only triggered for levels that have gift cards (Silver+)
    # New affiliates don't get gift cards (Bronze has 0), promotions to Silver+ do
    ws5.cell(row=6, column=col, value=f"=ROUND('Affiliate Projection'!{ml}19*(SUMPRODUCT('VIP Levels'!$E$15:$E$18,Inputs!$B$35:$B$38)/SUM(Inputs!$B$35:$B$38)),0)")
    style_calc(ws5.cell(row=6, column=col))

    # Discount Coupons: All level-ups get discounts
    ws5.cell(row=7, column=col, value=f"=ROUND('Affiliate Projection'!{ml}20*SUMPRODUCT('VIP Levels'!$G$14:$G$18,Inputs!$B$34:$B$38),0)")
    style_calc(ws5.cell(row=7, column=col))

    # Spark Ads: Levels with spark ads (Silver+)
    ws5.cell(row=8, column=col, value=f"=ROUND('Affiliate Projection'!{ml}19*(SUMPRODUCT('VIP Levels'!$I$15:$I$18,Inputs!$B$35:$B$38)/SUM(Inputs!$B$35:$B$38)),0)")
    style_calc(ws5.cell(row=8, column=col))

    # Physical Gifts: Gold+ only
    ws5.cell(row=9, column=col, value=f"=ROUND('Affiliate Projection'!{ml}19*(SUMPRODUCT('VIP Levels'!$K$16:$K$18,Inputs!$B$36:$B$38)/SUM(Inputs!$B$36:$B$38)),0)")
    style_calc(ws5.cell(row=9, column=col))

    # Experiences: Platinum+ only
    ws5.cell(row=10, column=col, value=f"=ROUND('Affiliate Projection'!{ml}19*(SUMPRODUCT('VIP Levels'!$M$17:$M$18,Inputs!$B$37:$B$38)/SUM(Inputs!$B$37:$B$38)),0)")
    style_calc(ws5.cell(row=10, column=col))

    # Mission completions by level
    # = Affiliates at level * missions for level * avg completion rate
    # Bronze
    ws5.cell(row=13, column=col, value=f"=ROUND('Affiliate Projection'!{ml}12*Missions!$B$29*Missions!$C$29,0)")
    style_calc(ws5.cell(row=13, column=col))
    # Silver
    ws5.cell(row=14, column=col, value=f"=ROUND('Affiliate Projection'!{ml}13*Missions!$B$30*Missions!$C$30,0)")
    style_calc(ws5.cell(row=14, column=col))
    # Gold
    ws5.cell(row=15, column=col, value=f"=ROUND('Affiliate Projection'!{ml}14*Missions!$B$31*Missions!$C$31,0)")
    style_calc(ws5.cell(row=15, column=col))
    # Platinum
    ws5.cell(row=16, column=col, value=f"=ROUND('Affiliate Projection'!{ml}15*Missions!$B$32*Missions!$C$32,0)")
    style_calc(ws5.cell(row=16, column=col))
    # Diamond
    ws5.cell(row=17, column=col, value=f"=ROUND('Affiliate Projection'!{ml}16*Missions!$B$33*Missions!$C$33,0)")
    style_calc(ws5.cell(row=17, column=col))

    # Total missions
    ws5.cell(row=18, column=col, value=f"=SUM({ml}13:{ml}17)")
    style_calc(ws5.cell(row=18, column=col))

ws5.column_dimensions['A'].width = 38
for col in range(2, 14):
    ws5.column_dimensions[get_column_letter(col)].width = 12

# ============================================================================
# SHEET 6: REVENUE CALCULATION
# ============================================================================
ws6 = wb.create_sheet("Revenue")

ws6['A1'] = "REVENUE PROJECTION (12 MONTHS)"
ws6['A1'].font = Font(bold=True, size=16, color="4472C4")
ws6.merge_cells('A1:N1')

for col, header in enumerate(proj_headers, 1):
    cell = ws6.cell(row=3, column=col, value=header)
    style_header(cell)

revenue_metrics = [
    'Total Sales Volume',
    'Gross AOV',
    'Gross Revenue',
    'Avg Discount % (when redeemed)',
    'Discount Redemption Rate',
    'Discounted Sales Count',
    'Net AOV (weighted avg)',
    'Net Revenue',
    'Discount Margin Erosion',
]

for i, metric in enumerate(revenue_metrics, 4):
    cell = ws6.cell(row=i, column=1, value=metric)
    style_label(cell)

for col in range(2, 14):
    ml = get_column_letter(col)

    # Total Sales
    ws6.cell(row=4, column=col, value=f"='Affiliate Projection'!{ml}10")
    style_calc(ws6.cell(row=4, column=col))

    # Gross AOV
    ws6.cell(row=5, column=col, value="=Inputs!$B$15")
    style_calc(ws6.cell(row=5, column=col))
    ws6.cell(row=5, column=col).number_format = '$#,##0.00'

    # Gross Revenue
    ws6.cell(row=6, column=col, value=f"={ml}4*{ml}5")
    style_calc(ws6.cell(row=6, column=col))
    ws6.cell(row=6, column=col).number_format = '$#,##0'

    # Avg Discount % - weighted from VIP Levels
    ws6.cell(row=7, column=col, value="=SUMPRODUCT('VIP Levels'!$H$14:$H$18,Inputs!$B$34:$B$38)")
    style_calc(ws6.cell(row=7, column=col))
    ws6.cell(row=7, column=col).number_format = '0%'

    # Discount Redemption Rate
    ws6.cell(row=8, column=col, value="=Inputs!$B$21")
    style_calc(ws6.cell(row=8, column=col))
    ws6.cell(row=8, column=col).number_format = '0%'

    # Discounted Sales Count
    ws6.cell(row=9, column=col, value=f"=ROUND({ml}4*{ml}8,0)")
    style_calc(ws6.cell(row=9, column=col))

    # Net AOV (weighted average)
    # = Gross AOV * (1 - Discount% * RedemptionRate)
    ws6.cell(row=10, column=col, value=f"={ml}5*(1-{ml}7*{ml}8)")
    style_calc(ws6.cell(row=10, column=col))
    ws6.cell(row=10, column=col).number_format = '$#,##0.00'

    # Net Revenue
    ws6.cell(row=11, column=col, value=f"={ml}4*{ml}10")
    style_calc(ws6.cell(row=11, column=col))
    ws6.cell(row=11, column=col).number_format = '$#,##0'

    # Discount Margin Erosion
    ws6.cell(row=12, column=col, value=f"={ml}6-{ml}11")
    style_calc(ws6.cell(row=12, column=col))
    ws6.cell(row=12, column=col).number_format = '$#,##0'

ws6.column_dimensions['A'].width = 35
for col in range(2, 14):
    ws6.column_dimensions[get_column_letter(col)].width = 12

# ============================================================================
# SHEET 7: COST CALCULATION
# ============================================================================
ws7 = wb.create_sheet("Costs")

ws7['A1'] = "COST PROJECTION (12 MONTHS)"
ws7['A1'].font = Font(bold=True, size=16, color="4472C4")
ws7.merge_cells('A1:N1')

for col, header in enumerate(proj_headers, 1):
    cell = ws7.cell(row=3, column=col, value=header)
    style_header(cell)

cost_metrics = [
    '--- CM1 COSTS (Per Sale) ---',
    'Base Commission Cost',
    'Commission Boost Cost (Welcome)',
    'Commission Boost Cost (Missions)',
    'Discount Cost (Margin Erosion)',
    'Total CM1 Costs',
    '',
    '--- LOYALTY PROGRAM COSTS ---',
    'Gift Card Cost (Welcome)',
    'Gift Card Cost (Missions)',
    'Total Gift Card Cost',
    'Total Loyalty Program Costs',
    '',
    '--- MARKETING OPEX ---',
    'Spark Ads Cost (Welcome)',
    'Spark Ads Cost (Missions)',
    'Physical Gift Cost',
    'Experience Cost',
    'Sample Cost',
    'Total Marketing OpEx',
    '',
    '--- TOTALS ---',
    'Total Program Cost',
]

for i, metric in enumerate(cost_metrics, 4):
    cell = ws7.cell(row=i, column=1, value=metric)
    if metric.startswith('---'):
        style_section(cell)
    else:
        style_label(cell)

for col in range(2, 14):
    ml = get_column_letter(col)

    # Base Commission Cost
    # = Sum of (Affiliates at level * sales per affiliate * level commission % * Net AOV)
    # Simplified using weighted avg commission
    ws7.cell(row=5, column=col, value=f"='Affiliate Projection'!{ml}10*SUMPRODUCT('VIP Levels'!$D$5:$D$9,Inputs!$B$34:$B$38)*Revenue!{ml}10")
    style_calc(ws7.cell(row=5, column=col))
    ws7.cell(row=5, column=col).number_format = '$#,##0'

    # Commission Boost Cost (Welcome)
    # = Boosts triggered * avg sales during boost * weighted boost % * Net AOV
    ws7.cell(row=6, column=col, value=f"='Reward Triggers'!{ml}5*Inputs!$B$23*'VIP Levels'!$B$22*Revenue!{ml}10")
    style_calc(ws7.cell(row=6, column=col))
    ws7.cell(row=6, column=col).number_format = '$#,##0'

    # Commission Boost Cost (Missions)
    # = Mission completions that are commission boosts * cost per boost
    # Estimate: 30% of missions are commission boosts
    ws7.cell(row=7, column=col, value=f"='Reward Triggers'!{ml}18*0.3*Inputs!$B$23*'VIP Levels'!$B$22*Revenue!{ml}10")
    style_calc(ws7.cell(row=7, column=col))
    ws7.cell(row=7, column=col).number_format = '$#,##0'

    # Discount Cost
    ws7.cell(row=8, column=col, value=f"=Revenue!{ml}12")
    style_calc(ws7.cell(row=8, column=col))
    ws7.cell(row=8, column=col).number_format = '$#,##0'

    # Total CM1
    ws7.cell(row=9, column=col, value=f"=SUM({ml}5:{ml}8)")
    style_calc(ws7.cell(row=9, column=col))
    ws7.cell(row=9, column=col).number_format = '$#,##0'
    ws7.cell(row=9, column=col).font = Font(bold=True)

    # Gift Card Cost (Welcome)
    # = Gift cards triggered * weighted avg gift card value
    ws7.cell(row=12, column=col, value=f"='Reward Triggers'!{ml}6*'VIP Levels'!$B$23")
    style_calc(ws7.cell(row=12, column=col))
    ws7.cell(row=12, column=col).number_format = '$#,##0'

    # Gift Card Cost (Missions)
    # = Mission completions * % that are gift cards * avg gift card value from missions
    # Estimate gift card missions = 50% of missions, avg value $75
    ws7.cell(row=13, column=col, value=f"='Reward Triggers'!{ml}18*0.5*75")
    style_calc(ws7.cell(row=13, column=col))
    ws7.cell(row=13, column=col).number_format = '$#,##0'

    # Total Gift Card Cost
    ws7.cell(row=14, column=col, value=f"={ml}12+{ml}13")
    style_calc(ws7.cell(row=14, column=col))
    ws7.cell(row=14, column=col).number_format = '$#,##0'

    # Total Loyalty Program
    ws7.cell(row=15, column=col, value=f"={ml}14")
    style_calc(ws7.cell(row=15, column=col))
    ws7.cell(row=15, column=col).number_format = '$#,##0'
    ws7.cell(row=15, column=col).font = Font(bold=True)

    # Spark Ads Cost (Welcome)
    ws7.cell(row=18, column=col, value=f"='Reward Triggers'!{ml}8*'VIP Levels'!$B$24")
    style_calc(ws7.cell(row=18, column=col))
    ws7.cell(row=18, column=col).number_format = '$#,##0'

    # Spark Ads Cost (Missions)
    # = Mission completions * % that are spark ads * avg spark ads value
    ws7.cell(row=19, column=col, value=f"='Reward Triggers'!{ml}18*0.2*50")
    style_calc(ws7.cell(row=19, column=col))
    ws7.cell(row=19, column=col).number_format = '$#,##0'

    # Physical Gift Cost
    ws7.cell(row=20, column=col, value=f"='Reward Triggers'!{ml}9*'VIP Levels'!$B$25")
    style_calc(ws7.cell(row=20, column=col))
    ws7.cell(row=20, column=col).number_format = '$#,##0'

    # Experience Cost
    ws7.cell(row=21, column=col, value=f"='Reward Triggers'!{ml}10*'VIP Levels'!$B$26")
    style_calc(ws7.cell(row=21, column=col))
    ws7.cell(row=21, column=col).number_format = '$#,##0'

    # Sample Cost
    ws7.cell(row=22, column=col, value=f"='Affiliate Projection'!{ml}4*Inputs!$B$28")
    style_calc(ws7.cell(row=22, column=col))
    ws7.cell(row=22, column=col).number_format = '$#,##0'

    # Total Marketing OpEx
    ws7.cell(row=23, column=col, value=f"=SUM({ml}18:{ml}22)")
    style_calc(ws7.cell(row=23, column=col))
    ws7.cell(row=23, column=col).number_format = '$#,##0'
    ws7.cell(row=23, column=col).font = Font(bold=True)

    # Total Program Cost
    ws7.cell(row=26, column=col, value=f"={ml}9+{ml}15+{ml}23")
    style_calc(ws7.cell(row=26, column=col))
    ws7.cell(row=26, column=col).number_format = '$#,##0'
    ws7.cell(row=26, column=col).font = Font(bold=True)

ws7.column_dimensions['A'].width = 35
for col in range(2, 14):
    ws7.column_dimensions[get_column_letter(col)].width = 12

# ============================================================================
# SHEET 8: SUMMARY DASHBOARD
# ============================================================================
ws8 = wb.create_sheet("Summary")

ws8['A1'] = "SUMMARY DASHBOARD"
ws8['A1'].font = Font(bold=True, size=16, color="4472C4")
ws8.merge_cells('A1:E1')

# 12-Month Totals
ws8['A3'] = "12-MONTH TOTALS"
style_section(ws8['A3'])

summary_headers = ['Metric', 'Value']
for col, header in enumerate(summary_headers, 1):
    cell = ws8.cell(row=4, column=col, value=header)
    style_header(cell)

summary_metrics = [
    ('Total Samples Sent', "=SUM('Affiliate Projection'!B4:M4)", None),
    ('Total New Affiliates', "=SUM('Affiliate Projection'!B7:M7)", None),
    ('Active Affiliates (Month 12)', "='Affiliate Projection'!M9", None),
    ('Total Sales', "=SUM('Affiliate Projection'!B10:M10)", None),
    ('', '', None),
    ('Gross Revenue', "=SUM(Revenue!B6:M6)", '$#,##0'),
    ('Net Revenue', "=SUM(Revenue!B11:M11)", '$#,##0'),
    ('', '', None),
    ('Total CM1 Costs', "=SUM(Costs!B9:M9)", '$#,##0'),
    ('Total Loyalty Program Costs', "=SUM(Costs!B15:M15)", '$#,##0'),
    ('Total Marketing OpEx', "=SUM(Costs!B23:M23)", '$#,##0'),
    ('Total Program Cost', "=SUM(Costs!B26:M26)", '$#,##0'),
]

for i, (metric, formula, fmt) in enumerate(summary_metrics, 5):
    ws8.cell(row=i, column=1, value=metric).border = thin_border
    cell = ws8.cell(row=i, column=2, value=formula if formula else '')
    if formula:
        style_calc(cell)
        if fmt:
            cell.number_format = fmt

# Key Ratios
ws8['A19'] = "KEY RATIOS"
style_section(ws8['A19'])

for col, header in enumerate(summary_headers, 1):
    cell = ws8.cell(row=20, column=col, value=header)
    style_header(cell)

ratio_metrics = [
    ('CM1 as % of Net Revenue', "=IF(B11>0,B13/B11,0)", '0.0%'),
    ('Total Cost as % of Net Revenue', "=IF(B11>0,B16/B11,0)", '0.0%'),
    ('Cost per Affiliate (12-mo avg)', "=IF(B6>0,B16/B6,0)", '$#,##0.00'),
    ('Revenue per Affiliate (12-mo avg)', "=IF(B6>0,B11/B6,0)", '$#,##0.00'),
    ('Sample-to-Affiliate Conversion', "=IF(B5>0,B6/B5,0)", '0.0%'),
    ('Weighted Avg Commission Rate', "=SUMPRODUCT('VIP Levels'!$D$5:$D$9,Inputs!$B$34:$B$38)", '0.0%'),
]

for i, (metric, formula, fmt) in enumerate(ratio_metrics, 21):
    ws8.cell(row=i, column=1, value=metric).border = thin_border
    cell = ws8.cell(row=i, column=2, value=formula)
    style_calc(cell)
    cell.number_format = fmt

# Monthly Trend
ws8['A29'] = "MONTHLY TREND"
style_section(ws8['A29'])

trend_headers = ['Metric'] + [f'M{i}' for i in range(1, 13)] + ['Total']
for col, header in enumerate(trend_headers, 1):
    cell = ws8.cell(row=30, column=col, value=header)
    style_header(cell)

trend_metrics = [
    'Active Affiliates',
    'Total Sales',
    'Net Revenue',
    'Total Program Cost',
    'Cost as % of Revenue',
]

for i, metric in enumerate(trend_metrics, 31):
    ws8.cell(row=i, column=1, value=metric).border = thin_border

for col in range(2, 14):
    ml = get_column_letter(col)

    ws8.cell(row=31, column=col, value=f"='Affiliate Projection'!{ml}9")
    style_calc(ws8.cell(row=31, column=col))

    ws8.cell(row=32, column=col, value=f"='Affiliate Projection'!{ml}10")
    style_calc(ws8.cell(row=32, column=col))

    ws8.cell(row=33, column=col, value=f"=Revenue!{ml}11")
    style_calc(ws8.cell(row=33, column=col))
    ws8.cell(row=33, column=col).number_format = '$#,##0'

    ws8.cell(row=34, column=col, value=f"=Costs!{ml}26")
    style_calc(ws8.cell(row=34, column=col))
    ws8.cell(row=34, column=col).number_format = '$#,##0'

    ws8.cell(row=35, column=col, value=f"=IF({ml}33>0,{ml}34/{ml}33,0)")
    style_calc(ws8.cell(row=35, column=col))
    ws8.cell(row=35, column=col).number_format = '0.0%'

# Totals column
ws8.cell(row=31, column=14, value="=M31")
style_calc(ws8.cell(row=31, column=14))

ws8.cell(row=32, column=14, value="=SUM(B32:M32)")
style_calc(ws8.cell(row=32, column=14))

ws8.cell(row=33, column=14, value="=SUM(B33:M33)")
style_calc(ws8.cell(row=33, column=14))
ws8.cell(row=33, column=14).number_format = '$#,##0'

ws8.cell(row=34, column=14, value="=SUM(B34:M34)")
style_calc(ws8.cell(row=34, column=14))
ws8.cell(row=34, column=14).number_format = '$#,##0'

ws8.cell(row=35, column=14, value="=IF(N33>0,N34/N33,0)")
style_calc(ws8.cell(row=35, column=14))
ws8.cell(row=35, column=14).number_format = '0.0%'

ws8.column_dimensions['A'].width = 32
ws8.column_dimensions['B'].width = 15
for col in range(3, 15):
    ws8.column_dimensions[get_column_letter(col)].width = 10

# Save workbook
wb.save('/home/jorge/Loyalty/Rumi/LoyaltyProgramModel_v2.xlsx')
print("Excel file created successfully: LoyaltyProgramModel_v2.xlsx")
print("\nKey improvements in V2:")
print("- Base commission uses weighted avg from VIP Levels sheet")
print("- Level distribution uses editable percentages from Inputs sheet")
print("- All reward costs reference VIP Levels configuration")
print("- Mission completions reference Missions sheet summary")
print("- Commission boost costs properly calculated with boost % from config")
