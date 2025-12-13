#!/usr/bin/env python3
"""
Loyalty Program Financial Model - Excel Generator
Creates a dynamic Excel workbook with formulas for modeling affiliate loyalty program costs.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, Reference

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
label_font = Font(bold=True, size=11)
section_font = Font(bold=True, size=14, color="4472C4")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(cell):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def style_input(cell):
    cell.fill = input_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

def style_calc(cell):
    cell.fill = calc_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

def style_label(cell):
    cell.font = label_font
    cell.border = thin_border

def style_section(cell):
    cell.font = section_font

# ============================================================================
# SHEET 1: INPUTS DASHBOARD
# ============================================================================
ws1 = wb.active
ws1.title = "Inputs"

# Title
ws1['A1'] = "LOYALTY PROGRAM FINANCIAL MODEL - INPUTS DASHBOARD"
ws1['A1'].font = Font(bold=True, size=16, color="4472C4")
ws1.merge_cells('A1:E1')

# Section: Acquisition Inputs
ws1['A3'] = "ACQUISITION INPUTS"
style_section(ws1['A3'])

headers = ['Parameter', 'Value', 'Unit', 'Description']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    style_header(cell)

acquisition_inputs = [
    ('Samples Sent - Month 1', 150, 'samples', 'Initial month sample distribution'),
    ('Samples Sent - Month 2+', 100, 'samples', 'Ongoing monthly sample distribution'),
    ('Sample-to-Affiliate Conversion Rate', 0.20, '%', 'Percentage of samples that convert to affiliates'),
    ('Flywheel Affiliates per Month', 10, 'affiliates', 'Organic inbound affiliates monthly'),
    ('Time to First Sale', 14, 'days', 'Lag between conversion and first sale'),
    ('Affiliate Attrition Rate', 0.05, '%', 'Monthly churn rate of affiliates'),
]

for i, (param, value, unit, desc) in enumerate(acquisition_inputs, 5):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    cell = ws1.cell(row=i, column=2, value=value)
    style_input(cell)
    if unit == '%':
        cell.number_format = '0%'
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=desc).border = thin_border

# Section: Performance Inputs
ws1['A12'] = "PERFORMANCE INPUTS"
style_section(ws1['A12'])

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=13, column=col, value=header)
    style_header(cell)

performance_inputs = [
    ('Average Sales per Affiliate per Month', 5, 'sales', 'Monthly sales velocity per active affiliate'),
    ('Gross AOV', 100, '$', 'Average order value before discounts'),
    ('Rolling Window Duration', 30, 'days', 'Time period for VIP level qualification'),
]

for i, (param, value, unit, desc) in enumerate(performance_inputs, 14):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    cell = ws1.cell(row=i, column=2, value=value)
    style_input(cell)
    if unit == '$':
        cell.number_format = '$#,##0.00'
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=desc).border = thin_border

# Section: Redemption Inputs
ws1['A19'] = "REDEMPTION & BEHAVIOR INPUTS"
style_section(ws1['A19'])

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=20, column=col, value=header)
    style_header(cell)

redemption_inputs = [
    ('Discount Redemption Rate', 0.30, '%', 'Percentage of discount coupons redeemed'),
    ('Mission Completion Rate (Default)', 0.25, '%', 'Default completion rate for missions'),
]

for i, (param, value, unit, desc) in enumerate(redemption_inputs, 21):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    cell = ws1.cell(row=i, column=2, value=value)
    style_input(cell)
    if unit == '%':
        cell.number_format = '0%'
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=desc).border = thin_border

# Section: Sample Cost Input
ws1['A25'] = "COST INPUTS"
style_section(ws1['A25'])

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=26, column=col, value=header)
    style_header(cell)

cost_inputs = [
    ('Cost per Sample', 15, '$', 'Cost of each sample sent (product + shipping)'),
]

for i, (param, value, unit, desc) in enumerate(cost_inputs, 27):
    ws1.cell(row=i, column=1, value=param).border = thin_border
    cell = ws1.cell(row=i, column=2, value=value)
    style_input(cell)
    if unit == '$':
        cell.number_format = '$#,##0.00'
    ws1.cell(row=i, column=3, value=unit).border = thin_border
    ws1.cell(row=i, column=4, value=desc).border = thin_border

# Legend
ws1['A30'] = "LEGEND"
style_section(ws1['A30'])
ws1['A31'] = "Yellow cells"
ws1['B31'] = "= Editable inputs"
ws1['B31'].fill = input_fill
ws1['A32'] = "Green cells"
ws1['B32'] = "= Calculated values"
ws1['B32'].fill = calc_fill

# Set column widths
ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 45

# Define named ranges for inputs
from openpyxl.workbook.defined_name import DefinedName

wb.defined_names['SamplesMonth1'] = DefinedName('SamplesMonth1', attr_text="'Inputs'!$B$5")
wb.defined_names['SamplesMonth2Plus'] = DefinedName('SamplesMonth2Plus', attr_text="'Inputs'!$B$6")
wb.defined_names['ConversionRate'] = DefinedName('ConversionRate', attr_text="'Inputs'!$B$7")
wb.defined_names['FlywheelAffiliates'] = DefinedName('FlywheelAffiliates', attr_text="'Inputs'!$B$8")
wb.defined_names['TimeToFirstSale'] = DefinedName('TimeToFirstSale', attr_text="'Inputs'!$B$9")
wb.defined_names['AttritionRate'] = DefinedName('AttritionRate', attr_text="'Inputs'!$B$10")
wb.defined_names['SalesPerAffiliate'] = DefinedName('SalesPerAffiliate', attr_text="'Inputs'!$B$14")
wb.defined_names['GrossAOV'] = DefinedName('GrossAOV', attr_text="'Inputs'!$B$15")
wb.defined_names['RollingWindow'] = DefinedName('RollingWindow', attr_text="'Inputs'!$B$16")
wb.defined_names['DiscountRedemptionRate'] = DefinedName('DiscountRedemptionRate', attr_text="'Inputs'!$B$21")
wb.defined_names['MissionCompletionRate'] = DefinedName('MissionCompletionRate', attr_text="'Inputs'!$B$22")
wb.defined_names['CostPerSample'] = DefinedName('CostPerSample', attr_text="'Inputs'!$B$27")

# ============================================================================
# SHEET 2: VIP LEVEL CONFIGURATION
# ============================================================================
ws2 = wb.create_sheet("VIP Levels")

ws2['A1'] = "VIP LEVEL CONFIGURATION"
ws2['A1'].font = Font(bold=True, size=16, color="4472C4")
ws2.merge_cells('A1:M1')

# Level Configuration Table
ws2['A3'] = "LEVEL THRESHOLDS & COMMISSION"
style_section(ws2['A3'])

level_headers = ['Level', 'Name', 'Sales Threshold', 'Base Commission %']
for col, header in enumerate(level_headers, 1):
    cell = ws2.cell(row=4, column=col, value=header)
    style_header(cell)

levels = [
    (1, 'Bronze', 0, 0.10),
    (2, 'Silver', 5, 0.12),
    (3, 'Gold', 20, 0.18),
    (4, 'Platinum', 50, 0.22),
    (5, 'Diamond', 100, 0.25),
]

for i, (level, name, threshold, commission) in enumerate(levels, 5):
    ws2.cell(row=i, column=1, value=level).border = thin_border
    cell = ws2.cell(row=i, column=2, value=name)
    style_input(cell)
    cell = ws2.cell(row=i, column=3, value=threshold)
    style_input(cell)
    cell = ws2.cell(row=i, column=4, value=commission)
    style_input(cell)
    cell.number_format = '0%'

# Welcome Rewards Configuration
ws2['A12'] = "WELCOME REWARDS (Per Level)"
style_section(ws2['A12'])

reward_headers = [
    'Level', 'Comm Boost Qty', 'Comm Boost %', 'Comm Boost Days',
    'Gift Card Qty', 'Gift Card $', 'Discount Qty', 'Discount %',
    'Spark Ads Qty', 'Spark Ads $', 'Physical Gift Qty', 'Physical Gift $',
    'Experience Qty', 'Experience $'
]
for col, header in enumerate(reward_headers, 1):
    cell = ws2.cell(row=13, column=col, value=header)
    style_header(cell)
    ws2.column_dimensions[get_column_letter(col)].width = 14

# Default welcome rewards per level
welcome_rewards = [
    # Level, CommBoostQty, CommBoost%, CommBoostDays, GiftCardQty, GiftCard$, DiscountQty, Discount%, SparkAdsQty, SparkAds$, PhysGiftQty, PhysGift$, ExpQty, Exp$
    ('Bronze', 1, 0.10, 30, 0, 0, 2, 0.15, 0, 0, 0, 0, 0, 0),
    ('Silver', 2, 0.15, 30, 1, 50, 2, 0.15, 1, 25, 0, 0, 0, 0),
    ('Gold', 1, 0.20, 30, 1, 100, 2, 0.20, 1, 50, 1, 50, 0, 0),
    ('Platinum', 2, 0.25, 45, 2, 150, 3, 0.25, 2, 75, 1, 75, 1, 100),
    ('Diamond', 3, 0.30, 60, 2, 200, 3, 0.30, 2, 100, 1, 100, 1, 200),
]

for i, rewards in enumerate(welcome_rewards, 14):
    for col, val in enumerate(rewards, 1):
        cell = ws2.cell(row=i, column=col, value=val)
        if col == 1:
            cell.border = thin_border
        else:
            style_input(cell)
        # Format percentages
        if col in [3, 8]:
            cell.number_format = '0%'
        # Format currency
        if col in [6, 10, 12, 14]:
            cell.number_format = '$#,##0'

# Named ranges for VIP levels
wb.defined_names['VIPLevels'] = DefinedName('VIPLevels', attr_text="'VIP Levels'!$B$5:$D$9")
wb.defined_names['WelcomeRewards'] = DefinedName('WelcomeRewards', attr_text="'VIP Levels'!$A$14:$N$18")

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 14

# ============================================================================
# SHEET 3: MISSION CONFIGURATION
# ============================================================================
ws3 = wb.create_sheet("Missions")

ws3['A1'] = "MISSION CONFIGURATION"
ws3['A1'].font = Font(bold=True, size=16, color="4472C4")
ws3.merge_cells('A1:K1')

ws3['A3'] = "MISSIONS BY VIP LEVEL"
style_section(ws3['A3'])

mission_headers = [
    'VIP Level', 'Mission Type', 'Target', 'Repeatability', 'Completion Rate',
    'Reward Type', 'Reward Value', 'Reward Duration (days)', 'Active?'
]
for col, header in enumerate(mission_headers, 1):
    cell = ws3.cell(row=4, column=col, value=header)
    style_header(cell)

# Sample missions
missions = [
    ('Bronze', 'Videos', 5, 'Monthly', 0.25, 'Gift Card', 50, 0, 'Yes'),
    ('Bronze', 'Likes', 100, 'Monthly', 0.20, 'Gift Card', 50, 0, 'Yes'),
    ('Silver', 'Videos', 10, 'Monthly', 0.20, 'Commission Boost', 0.20, 30, 'Yes'),
    ('Silver', 'Likes', 100, 'Monthly', 0.15, 'Gift Card', 50, 0, 'Yes'),
    ('Gold', 'Sales', 100, 'One-time', 0.10, 'Commission Boost', 0.30, 45, 'Yes'),
    ('Gold', 'Views', 1000, 'Monthly', 0.15, 'Spark Ads', 50, 0, 'Yes'),
    ('Platinum', 'Sales', 200, 'One-time', 0.08, 'Gift Card', 200, 0, 'Yes'),
    ('Platinum', 'Videos', 20, 'Monthly', 0.12, 'Commission Boost', 0.25, 30, 'Yes'),
    ('Diamond', 'Sales', 500, 'One-time', 0.05, 'Experience', 500, 0, 'Yes'),
    ('Diamond', 'Views', 5000, 'Monthly', 0.10, 'Spark Ads', 100, 0, 'Yes'),
]

for i, mission in enumerate(missions, 5):
    for col, val in enumerate(mission, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        if col in [4, 6, 9]:  # Text columns
            style_input(cell)
        elif col == 5:  # Completion rate
            style_input(cell)
            cell.number_format = '0%'
        elif col == 7:  # Reward value (could be % or $)
            style_input(cell)
            if mission[5] == 'Commission Boost':
                cell.number_format = '0%'
            else:
                cell.number_format = '$#,##0'
        else:
            style_input(cell)

# Add empty rows for more missions
for i in range(15, 25):
    for col in range(1, 10):
        cell = ws3.cell(row=i, column=col, value='')
        style_input(cell)

ws3['A26'] = "Mission Types: Videos, Likes, Sales, Views"
ws3['A27'] = "Repeatability Options: One-time, Weekly, Monthly"
ws3['A28'] = "Reward Types: Gift Card, Commission Boost, Spark Ads, Discount, Experience"

for col in range(1, 10):
    ws3.column_dimensions[get_column_letter(col)].width = 16

# ============================================================================
# SHEET 4: AFFILIATE PROJECTION
# ============================================================================
ws4 = wb.create_sheet("Affiliate Projection")

ws4['A1'] = "AFFILIATE PROJECTION (12 MONTHS)"
ws4['A1'].font = Font(bold=True, size=16, color="4472C4")
ws4.merge_cells('A1:N1')

# Headers
proj_headers = ['Metric'] + [f'Month {i}' for i in range(1, 13)]
for col, header in enumerate(proj_headers, 1):
    cell = ws4.cell(row=3, column=col, value=header)
    style_header(cell)

# Metrics
metrics = [
    'Samples Sent',
    'New Affiliates (from Samples)',
    'Flywheel Affiliates',
    'Total New Affiliates',
    'Churned Affiliates',
    'Active Affiliates (End of Month)',
    'Total Sales',
    'Affiliates at Bronze',
    'Affiliates at Silver',
    'Affiliates at Gold',
    'Affiliates at Platinum',
    'Affiliates at Diamond',
    'Level-Up Events',
    'Demotion Events',
]

for i, metric in enumerate(metrics, 4):
    cell = ws4.cell(row=i, column=1, value=metric)
    style_label(cell)

# Formulas for Month 1
# Row 4: Samples Sent
ws4['B4'] = "=Inputs!$B$5"
style_calc(ws4['B4'])

# Row 5: New Affiliates from Samples
ws4['B5'] = "=ROUND(B4*Inputs!$B$7,0)"
style_calc(ws4['B5'])

# Row 6: Flywheel Affiliates
ws4['B6'] = "=Inputs!$B$8"
style_calc(ws4['B6'])

# Row 7: Total New Affiliates
ws4['B7'] = "=B5+B6"
style_calc(ws4['B7'])

# Row 8: Churned Affiliates (0 for month 1)
ws4['B8'] = 0
style_calc(ws4['B8'])

# Row 9: Active Affiliates
ws4['B9'] = "=B7-B8"
style_calc(ws4['B9'])

# Row 10: Total Sales
ws4['B10'] = "=ROUND(B9*Inputs!$B$14,0)"
style_calc(ws4['B10'])

# Rows 11-15: Level distribution (simplified - starts all at Bronze)
ws4['B11'] = "=B9"  # Bronze
style_calc(ws4['B11'])
for row in range(12, 16):
    ws4.cell(row=row, column=2, value=0)
    style_calc(ws4.cell(row=row, column=2))

# Row 16: Level-Up Events
ws4['B16'] = "=B7"  # All new affiliates start at Bronze = level up
style_calc(ws4['B16'])

# Row 17: Demotion Events
ws4['B17'] = 0
style_calc(ws4['B17'])

# Formulas for Months 2-12
for col in range(3, 14):
    month_letter = get_column_letter(col)
    prev_letter = get_column_letter(col - 1)

    # Samples Sent (Month 2+)
    ws4.cell(row=4, column=col, value="=Inputs!$B$6")
    style_calc(ws4.cell(row=4, column=col))

    # New Affiliates from Samples
    ws4.cell(row=5, column=col, value=f"=ROUND({month_letter}4*Inputs!$B$7,0)")
    style_calc(ws4.cell(row=5, column=col))

    # Flywheel Affiliates
    ws4.cell(row=6, column=col, value="=Inputs!$B$8")
    style_calc(ws4.cell(row=6, column=col))

    # Total New Affiliates
    ws4.cell(row=7, column=col, value=f"={month_letter}5+{month_letter}6")
    style_calc(ws4.cell(row=7, column=col))

    # Churned Affiliates
    ws4.cell(row=8, column=col, value=f"=ROUND({prev_letter}9*Inputs!$B$10,0)")
    style_calc(ws4.cell(row=8, column=col))

    # Active Affiliates
    ws4.cell(row=9, column=col, value=f"={prev_letter}9+{month_letter}7-{month_letter}8")
    style_calc(ws4.cell(row=9, column=col))

    # Total Sales
    ws4.cell(row=10, column=col, value=f"=ROUND({month_letter}9*Inputs!$B$14,0)")
    style_calc(ws4.cell(row=10, column=col))

    # Level distribution (simplified model based on sales thresholds)
    # Bronze: Affiliates with < Silver threshold sales
    ws4.cell(row=11, column=col, value=f"=ROUND({month_letter}9*0.5,0)")  # Simplified: 50% stay Bronze
    style_calc(ws4.cell(row=11, column=col))

    # Silver
    ws4.cell(row=12, column=col, value=f"=ROUND({month_letter}9*0.25,0)")  # 25% reach Silver
    style_calc(ws4.cell(row=12, column=col))

    # Gold
    ws4.cell(row=13, column=col, value=f"=ROUND({month_letter}9*0.15,0)")  # 15% reach Gold
    style_calc(ws4.cell(row=13, column=col))

    # Platinum
    ws4.cell(row=14, column=col, value=f"=ROUND({month_letter}9*0.07,0)")  # 7% reach Platinum
    style_calc(ws4.cell(row=14, column=col))

    # Diamond
    ws4.cell(row=15, column=col, value=f"=ROUND({month_letter}9*0.03,0)")  # 3% reach Diamond
    style_calc(ws4.cell(row=15, column=col))

    # Level-Up Events (new affiliates + promotions estimate)
    ws4.cell(row=16, column=col, value=f"={month_letter}7+ROUND({prev_letter}9*0.1,0)")
    style_calc(ws4.cell(row=16, column=col))

    # Demotion Events
    ws4.cell(row=17, column=col, value=f"=ROUND({prev_letter}9*0.05,0)")
    style_calc(ws4.cell(row=17, column=col))

ws4.column_dimensions['A'].width = 35
for col in range(2, 14):
    ws4.column_dimensions[get_column_letter(col)].width = 12

# ============================================================================
# SHEET 5: REWARD TRIGGERS
# ============================================================================
ws5 = wb.create_sheet("Reward Triggers")

ws5['A1'] = "REWARD TRIGGERS (12 MONTHS)"
ws5['A1'].font = Font(bold=True, size=16, color="4472C4")
ws5.merge_cells('A1:N1')

# Headers
for col, header in enumerate(proj_headers, 1):
    cell = ws5.cell(row=3, column=col, value=header)
    style_header(cell)

reward_metrics = [
    '--- WELCOME REWARDS ---',
    'Commission Boosts Triggered',
    'Gift Cards Triggered (One-time)',
    'Discount Coupons Triggered',
    'Spark Ads Triggered',
    'Physical Gifts Triggered (One-time)',
    'Experiences Triggered (One-time)',
    '',
    '--- MISSION REWARDS ---',
    'Mission Completions (Bronze)',
    'Mission Completions (Silver)',
    'Mission Completions (Gold)',
    'Mission Completions (Platinum)',
    'Mission Completions (Diamond)',
    'Total Mission Completions',
]

for i, metric in enumerate(reward_metrics, 4):
    cell = ws5.cell(row=i, column=1, value=metric)
    if metric.startswith('---'):
        style_section(cell)
    else:
        style_label(cell)

# Formulas for each month
for col in range(2, 14):
    month_letter = get_column_letter(col)

    # Commission Boosts (triggered on level-up/re-qualification)
    # Sum of boosts per level * level-ups
    ws5.cell(row=5, column=col, value=f"='Affiliate Projection'!{month_letter}16*2")  # Avg 2 boosts per level-up
    style_calc(ws5.cell(row=5, column=col))

    # Gift Cards (one-time per level - only first time reaching level)
    ws5.cell(row=6, column=col, value=f"=ROUND('Affiliate Projection'!{month_letter}7*0.3,0)")  # 30% of new affiliates reach gift card levels
    style_calc(ws5.cell(row=6, column=col))

    # Discount Coupons
    ws5.cell(row=7, column=col, value=f"='Affiliate Projection'!{month_letter}16*2")
    style_calc(ws5.cell(row=7, column=col))

    # Spark Ads
    ws5.cell(row=8, column=col, value=f"=ROUND('Affiliate Projection'!{month_letter}16*0.5,0)")
    style_calc(ws5.cell(row=8, column=col))

    # Physical Gifts (one-time, higher levels only)
    ws5.cell(row=9, column=col, value=f"=ROUND('Affiliate Projection'!{month_letter}7*0.1,0)")
    style_calc(ws5.cell(row=9, column=col))

    # Experiences (one-time, highest levels only)
    ws5.cell(row=10, column=col, value=f"=ROUND('Affiliate Projection'!{month_letter}7*0.03,0)")
    style_calc(ws5.cell(row=10, column=col))

    # Mission completions per level (based on completion rate and active affiliates)
    ws5.cell(row=13, column=col, value=f"=ROUND('Affiliate Projection'!{month_letter}11*Inputs!$B$22*2,0)")  # Bronze missions
    style_calc(ws5.cell(row=13, column=col))

    ws5.cell(row=14, column=col, value=f"=ROUND('Affiliate Projection'!{month_letter}12*Inputs!$B$22*2,0)")  # Silver missions
    style_calc(ws5.cell(row=14, column=col))

    ws5.cell(row=15, column=col, value=f"=ROUND('Affiliate Projection'!{month_letter}13*Inputs!$B$22*2,0)")  # Gold missions
    style_calc(ws5.cell(row=15, column=col))

    ws5.cell(row=16, column=col, value=f"=ROUND('Affiliate Projection'!{month_letter}14*Inputs!$B$22*2,0)")  # Platinum missions
    style_calc(ws5.cell(row=16, column=col))

    ws5.cell(row=17, column=col, value=f"=ROUND('Affiliate Projection'!{month_letter}15*Inputs!$B$22*2,0)")  # Diamond missions
    style_calc(ws5.cell(row=17, column=col))

    # Total Mission Completions
    ws5.cell(row=18, column=col, value=f"=SUM({month_letter}13:{month_letter}17)")
    style_calc(ws5.cell(row=18, column=col))

ws5.column_dimensions['A'].width = 35
for col in range(2, 14):
    ws5.column_dimensions[get_column_letter(col)].width = 12

# ============================================================================
# SHEET 6: REVENUE CALCULATION
# ============================================================================
ws6 = wb.create_sheet("Revenue")

ws6['A1'] = "REVENUE PROJECTION (12 MONTHS)"
ws6['A1'].font = Font(bold=True, size=16, color="4472C4")
ws6.merge_cells('A1:N1')

for col, header in enumerate(proj_headers, 1):
    cell = ws6.cell(row=3, column=col, value=header)
    style_header(cell)

revenue_metrics = [
    'Total Sales Volume',
    'Gross AOV',
    'Gross Revenue',
    'Discount Rate Applied',
    'Discounted Sales',
    'Net AOV (avg)',
    'Net Revenue',
    'Gross Margin Erosion from Discounts',
]

for i, metric in enumerate(revenue_metrics, 4):
    cell = ws6.cell(row=i, column=1, value=metric)
    style_label(cell)

for col in range(2, 14):
    month_letter = get_column_letter(col)

    # Total Sales Volume
    ws6.cell(row=4, column=col, value=f"='Affiliate Projection'!{month_letter}10")
    style_calc(ws6.cell(row=4, column=col))

    # Gross AOV
    ws6.cell(row=5, column=col, value="=Inputs!$B$15")
    style_calc(ws6.cell(row=5, column=col))
    ws6.cell(row=5, column=col).number_format = '$#,##0.00'

    # Gross Revenue
    ws6.cell(row=6, column=col, value=f"={month_letter}4*{month_letter}5")
    style_calc(ws6.cell(row=6, column=col))
    ws6.cell(row=6, column=col).number_format = '$#,##0'

    # Discount Rate Applied (simplified: assume 15% avg discount on redeemed)
    ws6.cell(row=7, column=col, value=0.15)
    style_input(ws6.cell(row=7, column=col))
    ws6.cell(row=7, column=col).number_format = '0%'

    # Discounted Sales (sales using discount coupons)
    ws6.cell(row=8, column=col, value=f"=ROUND({month_letter}4*Inputs!$B$21,0)")
    style_calc(ws6.cell(row=8, column=col))

    # Net AOV (weighted average)
    ws6.cell(row=9, column=col, value=f"={month_letter}5*(1-{month_letter}7*Inputs!$B$21)")
    style_calc(ws6.cell(row=9, column=col))
    ws6.cell(row=9, column=col).number_format = '$#,##0.00'

    # Net Revenue
    ws6.cell(row=10, column=col, value=f"={month_letter}4*{month_letter}9")
    style_calc(ws6.cell(row=10, column=col))
    ws6.cell(row=10, column=col).number_format = '$#,##0'

    # Gross Margin Erosion
    ws6.cell(row=11, column=col, value=f"={month_letter}6-{month_letter}10")
    style_calc(ws6.cell(row=11, column=col))
    ws6.cell(row=11, column=col).number_format = '$#,##0'

ws6.column_dimensions['A'].width = 35
for col in range(2, 14):
    ws6.column_dimensions[get_column_letter(col)].width = 12

# ============================================================================
# SHEET 7: COST CALCULATION
# ============================================================================
ws7 = wb.create_sheet("Costs")

ws7['A1'] = "COST PROJECTION (12 MONTHS)"
ws7['A1'].font = Font(bold=True, size=16, color="4472C4")
ws7.merge_cells('A1:N1')

for col, header in enumerate(proj_headers, 1):
    cell = ws7.cell(row=3, column=col, value=header)
    style_header(cell)

cost_metrics = [
    '--- CM1 COSTS (Per Sale) ---',
    'Base Commission Cost',
    'Commission Boost Cost',
    'Discount Cost (Margin Erosion)',
    'Total CM1 Costs',
    '',
    '--- LOYALTY PROGRAM COSTS ---',
    'Gift Card Cost',
    'Total Loyalty Program Costs',
    '',
    '--- MARKETING OPEX ---',
    'Spark Ads Cost',
    'Physical Gift Cost',
    'Experience Cost',
    'Sample Cost',
    'Total Marketing OpEx',
    '',
    '--- TOTALS ---',
    'Total Program Cost',
]

for i, metric in enumerate(cost_metrics, 4):
    cell = ws7.cell(row=i, column=1, value=metric)
    if metric.startswith('---'):
        style_section(cell)
    else:
        style_label(cell)

for col in range(2, 14):
    month_letter = get_column_letter(col)

    # Base Commission Cost (sales * avg commission rate * Net AOV)
    # Weighted avg commission ~12%
    ws7.cell(row=5, column=col, value=f"='Affiliate Projection'!{month_letter}10*0.12*Revenue!{month_letter}9")
    style_calc(ws7.cell(row=5, column=col))
    ws7.cell(row=5, column=col).number_format = '$#,##0'

    # Commission Boost Cost
    # Boost triggers * avg sales during boost * boost % * Net AOV
    ws7.cell(row=6, column=col, value=f"='Reward Triggers'!{month_letter}5*3*0.15*Revenue!{month_letter}9")
    style_calc(ws7.cell(row=6, column=col))
    ws7.cell(row=6, column=col).number_format = '$#,##0'

    # Discount Cost (margin erosion)
    ws7.cell(row=7, column=col, value=f"=Revenue!{month_letter}11")
    style_calc(ws7.cell(row=7, column=col))
    ws7.cell(row=7, column=col).number_format = '$#,##0'

    # Total CM1
    ws7.cell(row=8, column=col, value=f"=SUM({month_letter}5:{month_letter}7)")
    style_calc(ws7.cell(row=8, column=col))
    ws7.cell(row=8, column=col).number_format = '$#,##0'
    ws7.cell(row=8, column=col).font = Font(bold=True)

    # Gift Card Cost (triggers * avg value $75)
    ws7.cell(row=11, column=col, value=f"='Reward Triggers'!{month_letter}6*75+'Reward Triggers'!{month_letter}18*50*0.3")
    style_calc(ws7.cell(row=11, column=col))
    ws7.cell(row=11, column=col).number_format = '$#,##0'

    # Total Loyalty Program
    ws7.cell(row=12, column=col, value=f"={month_letter}11")
    style_calc(ws7.cell(row=12, column=col))
    ws7.cell(row=12, column=col).number_format = '$#,##0'
    ws7.cell(row=12, column=col).font = Font(bold=True)

    # Spark Ads Cost
    ws7.cell(row=15, column=col, value=f"='Reward Triggers'!{month_letter}8*50+'Reward Triggers'!{month_letter}18*25*0.2")
    style_calc(ws7.cell(row=15, column=col))
    ws7.cell(row=15, column=col).number_format = '$#,##0'

    # Physical Gift Cost
    ws7.cell(row=16, column=col, value=f"='Reward Triggers'!{month_letter}9*60")
    style_calc(ws7.cell(row=16, column=col))
    ws7.cell(row=16, column=col).number_format = '$#,##0'

    # Experience Cost
    ws7.cell(row=17, column=col, value=f"='Reward Triggers'!{month_letter}10*150")
    style_calc(ws7.cell(row=17, column=col))
    ws7.cell(row=17, column=col).number_format = '$#,##0'

    # Sample Cost
    ws7.cell(row=18, column=col, value=f"='Affiliate Projection'!{month_letter}4*Inputs!$B$27")
    style_calc(ws7.cell(row=18, column=col))
    ws7.cell(row=18, column=col).number_format = '$#,##0'

    # Total Marketing OpEx
    ws7.cell(row=19, column=col, value=f"=SUM({month_letter}15:{month_letter}18)")
    style_calc(ws7.cell(row=19, column=col))
    ws7.cell(row=19, column=col).number_format = '$#,##0'
    ws7.cell(row=19, column=col).font = Font(bold=True)

    # Total Program Cost
    ws7.cell(row=22, column=col, value=f"={month_letter}8+{month_letter}12+{month_letter}19")
    style_calc(ws7.cell(row=22, column=col))
    ws7.cell(row=22, column=col).number_format = '$#,##0'
    ws7.cell(row=22, column=col).font = Font(bold=True)

ws7.column_dimensions['A'].width = 35
for col in range(2, 14):
    ws7.column_dimensions[get_column_letter(col)].width = 12

# ============================================================================
# SHEET 8: SUMMARY DASHBOARD
# ============================================================================
ws8 = wb.create_sheet("Summary")

ws8['A1'] = "SUMMARY DASHBOARD"
ws8['A1'].font = Font(bold=True, size=16, color="4472C4")
ws8.merge_cells('A1:E1')

# 12-Month Totals
ws8['A3'] = "12-MONTH TOTALS"
style_section(ws8['A3'])

summary_headers = ['Metric', 'Value']
for col, header in enumerate(summary_headers, 1):
    cell = ws8.cell(row=4, column=col, value=header)
    style_header(cell)

summary_metrics = [
    ('Total Samples Sent', "=SUM('Affiliate Projection'!B4:M4)"),
    ('Total New Affiliates', "=SUM('Affiliate Projection'!B7:M7)"),
    ('Active Affiliates (Month 12)', "='Affiliate Projection'!M9"),
    ('Total Sales', "=SUM('Affiliate Projection'!B10:M10)"),
    ('', ''),
    ('Gross Revenue', "=SUM(Revenue!B6:M6)"),
    ('Net Revenue', "=SUM(Revenue!B10:M10)"),
    ('', ''),
    ('Total CM1 Costs', "=SUM(Costs!B8:M8)"),
    ('Total Loyalty Program Costs', "=SUM(Costs!B12:M12)"),
    ('Total Marketing OpEx', "=SUM(Costs!B19:M19)"),
    ('Total Program Cost', "=SUM(Costs!B22:M22)"),
]

for i, (metric, formula) in enumerate(summary_metrics, 5):
    ws8.cell(row=i, column=1, value=metric).border = thin_border
    cell = ws8.cell(row=i, column=2, value=formula if formula else '')
    if formula:
        style_calc(cell)
        if 'Revenue' in formula or 'Costs' in formula:
            cell.number_format = '$#,##0'

# Key Ratios
ws8['A19'] = "KEY RATIOS"
style_section(ws8['A19'])

for col, header in enumerate(summary_headers, 1):
    cell = ws8.cell(row=20, column=col, value=header)
    style_header(cell)

ratio_metrics = [
    ('CM1 as % of Net Revenue', "=B13/B11"),
    ('Total Cost as % of Net Revenue', "=B16/B11"),
    ('Cost per Affiliate (12-mo)', "=B16/B6"),
    ('Revenue per Affiliate (12-mo)', "=B11/B6"),
    ('Sample-to-Affiliate Conversion', "=B6/B5"),
]

for i, (metric, formula) in enumerate(ratio_metrics, 21):
    ws8.cell(row=i, column=1, value=metric).border = thin_border
    cell = ws8.cell(row=i, column=2, value=formula)
    style_calc(cell)
    if 'per Affiliate' in metric or 'Revenue per' in metric:
        cell.number_format = '$#,##0.00'
    else:
        cell.number_format = '0.0%'

# Monthly Trend Headers
ws8['A28'] = "MONTHLY TREND"
style_section(ws8['A28'])

trend_headers = ['Metric'] + [f'M{i}' for i in range(1, 13)] + ['Total']
for col, header in enumerate(trend_headers, 1):
    cell = ws8.cell(row=29, column=col, value=header)
    style_header(cell)

trend_metrics = [
    'Active Affiliates',
    'Total Sales',
    'Net Revenue',
    'Total Program Cost',
    'Cost as % of Revenue',
]

for i, metric in enumerate(trend_metrics, 30):
    ws8.cell(row=i, column=1, value=metric).border = thin_border

# Fill in monthly data with formulas
for col in range(2, 14):
    month_letter = get_column_letter(col)

    # Active Affiliates
    ws8.cell(row=30, column=col, value=f"='Affiliate Projection'!{month_letter}9")
    style_calc(ws8.cell(row=30, column=col))

    # Total Sales
    ws8.cell(row=31, column=col, value=f"='Affiliate Projection'!{month_letter}10")
    style_calc(ws8.cell(row=31, column=col))

    # Net Revenue
    ws8.cell(row=32, column=col, value=f"=Revenue!{month_letter}10")
    style_calc(ws8.cell(row=32, column=col))
    ws8.cell(row=32, column=col).number_format = '$#,##0'

    # Total Program Cost
    ws8.cell(row=33, column=col, value=f"=Costs!{month_letter}22")
    style_calc(ws8.cell(row=33, column=col))
    ws8.cell(row=33, column=col).number_format = '$#,##0'

    # Cost as % of Revenue
    ws8.cell(row=34, column=col, value=f"=IF({month_letter}32>0,{month_letter}33/{month_letter}32,0)")
    style_calc(ws8.cell(row=34, column=col))
    ws8.cell(row=34, column=col).number_format = '0.0%'

# Totals column
ws8.cell(row=30, column=14, value="=M30")  # End of period affiliates
style_calc(ws8.cell(row=30, column=14))

ws8.cell(row=31, column=14, value="=SUM(B31:M31)")
style_calc(ws8.cell(row=31, column=14))

ws8.cell(row=32, column=14, value="=SUM(B32:M32)")
style_calc(ws8.cell(row=32, column=14))
ws8.cell(row=32, column=14).number_format = '$#,##0'

ws8.cell(row=33, column=14, value="=SUM(B33:M33)")
style_calc(ws8.cell(row=33, column=14))
ws8.cell(row=33, column=14).number_format = '$#,##0'

ws8.cell(row=34, column=14, value="=IF(N32>0,N33/N32,0)")
style_calc(ws8.cell(row=34, column=14))
ws8.cell(row=34, column=14).number_format = '0.0%'

ws8.column_dimensions['A'].width = 30
ws8.column_dimensions['B'].width = 15
for col in range(3, 15):
    ws8.column_dimensions[get_column_letter(col)].width = 10

# Save workbook
wb.save('/home/jorge/Loyalty/Rumi/LoyaltyProgramModel.xlsx')
print("Excel file created successfully: LoyaltyProgramModel.xlsx")
